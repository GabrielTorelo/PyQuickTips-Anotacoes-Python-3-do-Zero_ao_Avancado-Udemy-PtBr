# Importa o módulo openpyxl (usado para manipular planilhas)
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Cria uma nova planilha
wb = Workbook()

# Ativa a planilha
ws: Worksheet = wb.active

# Adiciona cabeçalho para as colunas
ws['A1'] = 'Nome'
ws['B1'] = 'Idade'

# Adiciona valores nas células
ws['A2'] = 'Gabriel'
ws['B2'] = '24'

# Salva a planilha
wb.save('aula115.xlsx')

# --------------------------------------------
# --------------------------------------------

# Carrega um workbook existente
load_wb = load_workbook('aula115.xlsx')

# Seleciona a planilha ativa
load_ws: Worksheet = load_wb.active

print('Celulas da primeira linha:')

# Exibe os valores das células da segunda linha
for cell in load_ws['1:1']:
  print(f'{cell.column}{cell.row}: {cell.value}')

# Exibe os valores das células A2 e B2
print(f'\nValor da célula A2: {load_ws["A2"].value}')
print(f'Valor da célula B2: {load_ws["B2"].value}')
