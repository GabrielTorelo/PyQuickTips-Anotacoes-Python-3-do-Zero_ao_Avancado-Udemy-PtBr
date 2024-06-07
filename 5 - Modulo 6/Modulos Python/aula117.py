# Importa o módulo json (usado para manipular arquivos JSON)
from json import load

# Importa o módulo openpyxl (usado para manipular planilhas)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet

# Cria uma nova planilha
wb = Workbook()

# Ativa a planilha
ws: Worksheet = wb.active

# Cria nome da planilha
planilha_nome = 'Notas da Turma'

# Renomeia a planilha
wb.active.title = planilha_nome

# Adiciona um cabeçalho
ws.append(['Aluno', 'Nota', 'Situação'])

# Cria um objeto de estilo para as células (cor de fundo)
celula_preta = PatternFill(start_color='000000', fill_type = 'solid')
celula_vermelha = PatternFill(start_color='FF0000', fill_type = 'solid')
celula_verde = PatternFill(start_color='00FF00', fill_type = 'solid')

# Cria um objeto de estilo para as células (cor da fonte)
fonte_branca = Font(color='FFFFFF')

# Cria um objeto de estilo para as células (bordas)
borda = Border(
  left=Side(border_style='thin', color='000000'),
  right=Side(border_style='thin', color='000000'),
  top=Side(border_style='thin', color='000000'),
  bottom=Side(border_style='thin', color='000000')
)

# Cria um objeto de estilo para as células (alinhamento)
alinhamento_centro = Alignment(horizontal='center', vertical='center')

# Altera o estilo das células
  # Cor da fonte
ws['A1'].font = fonte_branca
ws['B1'].font = fonte_branca
ws['C1'].font = fonte_branca
  # Cor de fundo
ws['A1'].fill = celula_preta
ws['B1'].fill = celula_preta
ws['C1'].fill = celula_preta
  # Alinhamento
ws['A1'].alignment = alinhamento_centro
ws['B1'].alignment = alinhamento_centro
ws['C1'].alignment = alinhamento_centro

# Cria uma lista vazia para armazenar os estudantes
estudantes = []

# Pega os dados dos estudantes do arquivo JSON
try:
  print('Lendo arquivo JSON com os dados dos estudantes! Por favor, aguarde...')

  # Abre o arquivo JSON em modo de leitura
  with open('aula117.json', 'r', encoding='utf8') as arquivo:

    # Carrega os dados do arquivo JSON
    for estudante in load(arquivo):

      # Adiciona os dados dos estudantes na lista
      estudantes.append([estudante['Nome'], estudante['Nota']])

  print('\nArquivo JSON lido com sucesso!')
except Exception as e: # Trata caso ocorra algum erro
  print('\nErro ao abrir o arquivo JSON:', e)

print('\nAdicionando os estudantes na planilha! Por favor, aguarde...')

# Adiciona os estudantes na planilha
for estudante in estudantes:
  ws.append(estudante)

  # Adiciona bordas nas células
  ws[f'A{ws.max_row}'].border = borda
  ws[f'B{ws.max_row}'].border = borda
  ws[f'C{ws.max_row}'].border = borda

  # Centraliza o texto das células
  ws[f'A{ws.max_row}'].alignment = alinhamento_centro
  ws[f'B{ws.max_row}'].alignment = alinhamento_centro
  ws[f'C{ws.max_row}'].alignment = alinhamento_centro

  # Verifica se a nota é menor que 6
  if estudante[1] < 6:

    # Adiciona cor vermelha nas células da coluna B e C
    ws[f'B{ws.max_row}'].fill = celula_vermelha
    ws[f'C{ws.max_row}'].fill = celula_vermelha

    # Adiciona a palavra 'Reprovado' na célula da coluna C
    ws[f'C{ws.max_row}'] = 'Reprovado'
    
    # Adiciona fonte branca nas células da coluna C
    ws[f'C{ws.max_row}'].font = fonte_branca
  else: # Caso contrário

    # Adiciona cor verde nas células da coluna B e C
    ws[f'B{ws.max_row}'].fill = celula_verde
    ws[f'C{ws.max_row}'].fill = celula_verde

    # Adiciona a palavra 'Aprovado' na célula da coluna C
    ws[f'C{ws.max_row}'] = 'Aprovado'

# Salva a planilha
wb.save('aula117.xlsx')
