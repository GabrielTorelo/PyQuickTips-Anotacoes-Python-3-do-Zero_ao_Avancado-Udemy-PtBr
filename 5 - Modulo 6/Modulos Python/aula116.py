# Importa o módulo openpyxl (usado para manipular planilhas)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# Carrega um workbook existente
load_wb = load_workbook('aula115.xlsx')

# Seleciona a planilha ativa
load_ws: Worksheet = load_wb.active

# Cria um objeto de estilo para as células (cor de fundo)
celula_vermelha = PatternFill(start_color="FF0000", fill_type = "solid")
celula_verde = PatternFill(start_color="00FF00", fill_type = "solid")

# Cria um objeto de estilo para as células (cor da fonte)
fonte_branca = Font(color='FFFFFF')

# Altera os valores das células
load_ws['A2'] = 'José'
load_ws['B2'] = '33'

# Altera o estilo das células
load_ws['A2'].font = fonte_branca
load_ws['B2'].font = fonte_branca
load_ws['A2'].fill = celula_vermelha
load_ws['B2'].fill = celula_verde

# Salva a planilha
load_wb.save('aula116.xlsx')
